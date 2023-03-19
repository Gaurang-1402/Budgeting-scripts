import pandas as pd
from openpyxl import load_workbook
import os

from dotenv import load_dotenv

load_dotenv()

# Load the workbooks
transactions_workbook = load_workbook(os.getenv("SRC_WB_PATH"))
budget_workbook = load_workbook(os.getenv("DEST_WB_PATH"))
# Get the sheets
transactions_sheet = transactions_workbook['Transactions']
budget_sheet = budget_workbook['Budget Tracking']

# Get the header row from 'Budget Tracking'
header_row = []
for cell in budget_sheet[11]:
    header_row.append(cell.value)

# Create dictionary mapping column headers in "Transactions" to corresponding column index in "Budget Tracking"
header_dict = {}
for col in range(3, budget_sheet.max_column+1):
    if budget_sheet.cell(row=11, column=col).value in header_row:
        header_dict[budget_sheet.cell(row=11, column=col).value] = col

# Get the rows to copy from 'Transactions'
rows_to_copy = []
for row in range(2, transactions_sheet.max_row+1):
    row_to_copy = []
    for header in header_row:
        if header in header_dict:
            col = header_dict[header]
            row_to_copy.append(transactions_sheet.cell(row=row, column=col).value)
        else:
            row_to_copy.append(None)
    rows_to_copy.append(row_to_copy)

print(header_dict)

print(rows_to_copy[-1])

# Write the rows to 'Budget Tracking'
for i, row in enumerate(rows_to_copy):
    for j, value in enumerate(row):
        budget_sheet.cell(row=i+12, column=j+3, value=value)

# Save the changes
budget_workbook.save(os.getenv("DEST_WB_PATH"))
