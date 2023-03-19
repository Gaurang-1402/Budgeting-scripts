import openpyxl

# Load the workbooks
transactions_wb = openpyxl.load_workbook('./my_test.xlsx')
playaround_wb = openpyxl.load_workbook('./tracking_test.xlsx')

# Get the "Transactions" and "Budget Tracking" sheets
transactions_ws = transactions_wb['Transactions']
playaround_ws = playaround_wb['Budget Tracking']

# Get the headers from the "Budget Tracking" sheet
headers = [cell.value.lower() if cell.value != None else None for cell in playaround_ws[11]]
print(headers)
# Loop through the rows in "Transactions" and copy the data to "Budget Tracking"
for row in transactions_ws.iter_rows(min_row=2, values_only=True):
    data = dict(zip(headers, row))
    
    if all(key in data for key in headers):
        playaround_ws.append([
            '', '', data['date'], '', data['category'],
            abs(float(data['amount'])), data['description']
        ])

# Save the changes to the "playaround" workbook
playaround_wb.save('playaround.xlsx')
